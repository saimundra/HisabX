import pandas as pd
import io
from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def generate_excel_report(report_data, report_title, user_name):
    """Generate Excel report from report data"""
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Expense Report"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = report_title
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # User and date info
    ws['A2'] = f"Generated for: {user_name}"
    ws['A3'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    # Summary section
    ws['A5'] = "SUMMARY"
    ws['A5'].font = header_font
    ws['A5'].fill = header_fill
    
    total_amount = sum(item.get('total_amount', 0) for item in report_data)
    total_bills = sum(item.get('bill_count', 0) for item in report_data)
    
    ws['A6'] = f"Total Bills: {total_bills}"
    ws['A7'] = f"Total Amount: ${total_amount:.2f}"
    
    # Category breakdown
    ws['A9'] = "CATEGORY BREAKDOWN"
    ws['A9'].font = header_font
    ws['A9'].fill = header_fill
    
    # Headers for category data
    headers = ['Category', 'Bills Count', 'Total Amount', 'Percentage']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Category data
    for row, item in enumerate(report_data, 11):
        ws.cell(row=row, column=1, value=item.get('category__name', 'Uncategorized'))
        ws.cell(row=row, column=2, value=item.get('bill_count', 0))
        ws.cell(row=row, column=3, value=f"${item.get('total_amount', 0):.2f}")
        percentage = (item.get('total_amount', 0) / total_amount * 100) if total_amount > 0 else 0
        ws.cell(row=row, column=4, value=f"{percentage:.1f}%")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

def create_csv_report(report_data, report_title):
    """Generate CSV report from report data"""
    
    # Create DataFrame
    df_data = []
    for item in report_data:
        df_data.append({
            'Category': item.get('category__name', 'Uncategorized'),
            'Bills Count': item.get('bill_count', 0),
            'Total Amount': item.get('total_amount', 0),
            'Percentage': (item.get('total_amount', 0) / sum(i.get('total_amount', 0) for i in report_data) * 100) if report_data else 0
        })
    
    df = pd.DataFrame(df_data)
    
    # Add summary row
    total_bills = df['Bills Count'].sum()
    total_amount = df['Total Amount'].sum()
    
    summary_row = pd.DataFrame([{
        'Category': 'TOTAL',
        'Bills Count': total_bills,
        'Total Amount': total_amount,
        'Percentage': 100.0
    }])
    
    df = pd.concat([df, summary_row], ignore_index=True)
    
    # Convert to CSV
    output = io.StringIO()
    df.to_csv(output, index=False)
    output.seek(0)
    
    return output.getvalue()

def export_report_response(report_data, format_type, filename, report_title, user_name):
    """Create HTTP response for report export"""
    
    if format_type == 'EXCEL':
        excel_file = generate_excel_report(report_data, report_title, user_name)
        response = HttpResponse(
            excel_file.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        return response
    
    elif format_type == 'CSV':
        csv_content = create_csv_report(report_data, report_title)
        response = HttpResponse(csv_content, content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        return response
    
    else:
        raise ValueError(f"Unsupported format: {format_type}")