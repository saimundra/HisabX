"""
Report Export Service for PDF and Excel
Generates downloadable financial reports for auditing
"""
from io import BytesIO
from datetime import datetime
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_RIGHT


class ReportExporter:
    """Export financial reports to Excel and PDF"""
    
    def __init__(self, user):
        self.user = user
    
    def export_balance_sheet_excel(self, balance_sheet_data):
        """
        Export Balance Sheet to Excel
        
        Args:
            balance_sheet_data: Dict from FinancialStatementsService.get_balance_sheet()
        
        Returns:
            HttpResponse with Excel file
        """
        import logging
        logger = logging.getLogger(__name__)
        
        logger.info(f"=== export_balance_sheet_excel called ===")
        logger.info(f"Method: export_balance_sheet_excel (Balance Sheet)")
        logger.info(f"User: {self.user.username}")
        
        # Validate input data
        if not balance_sheet_data:
            raise ValueError("Balance sheet data is required")
        
        if 'as_of_date' not in balance_sheet_data:
            raise ValueError("Balance sheet data must include 'as_of_date'")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Balance Sheet"
        
        # Title
        user_name = self.user.get_full_name() or self.user.username
        ws['A1'] = f"{user_name} - Balance Sheet"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"As of {balance_sheet_data['as_of_date']}"
        ws['A2'].font = Font(size=12, italic=True)
        
        # Headers
        row = 4
        ws[f'A{row}'] = "Account"
        ws[f'B{row}'] = "Amount"
        for col in ['A', 'B']:
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{col}{row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws[f'{col}{row}'].font = Font(color="FFFFFF", bold=True)
        
        row += 1
        
        # ASSETS
        ws[f'A{row}'] = "ASSETS"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        
        # Validate required data structure
        assets = balance_sheet_data.get('assets', {})
        liabilities = balance_sheet_data.get('liabilities', {})
        equity = balance_sheet_data.get('equity', {})
        
        ws[f'A{row}'] = "Current Assets"
        ws[f'B{row}'] = float(assets.get('current_assets', 0))
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 1
        
        ws[f'A{row}'] = "Total Assets"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = float(assets.get('total_assets', 0))
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'B{row}'].font = Font(bold=True)
        row += 2
        
        # LIABILITIES
        ws[f'A{row}'] = "LIABILITIES"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        
        ws[f'A{row}'] = "Current Liabilities"
        ws[f'B{row}'] = float(liabilities.get('current_liabilities', 0))
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 1
        
        ws[f'A{row}'] = "Total Liabilities"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = float(liabilities.get('total_liabilities', 0))
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'B{row}'].font = Font(bold=True)
        row += 2
        
        # EQUITY
        ws[f'A{row}'] = "EQUITY"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        
        ws[f'A{row}'] = "Retained Earnings"
        ws[f'B{row}'] = float(equity.get('retained_earnings', 0))
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 1
        
        ws[f'A{row}'] = "Other Equity"
        ws[f'B{row}'] = float(equity.get('other_equity', 0))
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 1
        
        ws[f'A{row}'] = "Total Equity"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = float(equity.get('total_equity', 0))
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'B{row}'].font = Font(bold=True)
        row += 2
        
        ws[f'A{row}'] = "Total Liabilities & Equity"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = float(balance_sheet_data.get('total_liabilities_and_equity', 0))
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'B{row}'].font = Font(bold=True)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        file_content = output.read()
        output.close()
        
        # Ensure filename is safe
        as_of_date = balance_sheet_data.get('as_of_date', 'unknown')
        if isinstance(as_of_date, str):
            filename = f"balance_sheet_{as_of_date}.xlsx"
        else:
            filename = f"balance_sheet_{as_of_date}.xlsx"
        
        response = HttpResponse(
            file_content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Content-Length'] = len(file_content)
        return response
    
    def export_income_statement_excel(self, income_data):
        """Export Income Statement (P&L) to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Income Statement"
        
        # Title
        ws['A1'] = f"{self.user.get_full_name() or self.user.username} - Income Statement"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Period: {income_data['period']['start_date']} to {income_data['period']['end_date']}"
        ws['A2'].font = Font(size=12, italic=True)
        
        # Headers
        row = 4
        ws[f'A{row}'] = "Description"
        ws[f'B{row}'] = "Amount"
        for col in ['A', 'B']:
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{col}{row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            ws[f'{col}{row}'].font = Font(color="FFFFFF", bold=True)
        
        row += 1
        
        # Revenue
        ws[f'A{row}'] = "REVENUE"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        
        ws[f'A{row}'] = "Total Revenue"
        ws[f'B{row}'] = float(income_data['revenue'])
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'B{row}'].font = Font(bold=True)
        row += 2
        
        # Expenses
        ws[f'A{row}'] = "EXPENSES"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        
        for category, amount in income_data['expenses']['breakdown'].items():
            ws[f'A{row}'] = category
            ws[f'B{row}'] = float(amount)
            ws[f'B{row}'].number_format = '#,##0.00'
            row += 1
        
        ws[f'A{row}'] = "Total Expenses"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = float(income_data['expenses']['total_expenses'])
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'B{row}'].font = Font(bold=True)
        row += 2
        
        # Net Income
        ws[f'A{row}'] = "NET INCOME"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'B{row}'] = float(income_data['net_income'])
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'B{row}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        row += 1
        
        ws[f'A{row}'] = "Profit Margin"
        ws[f'B{row}'] = f"{income_data['profit_margin']:.2f}%"
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'income_statement_{income_data["period"]["start_date"]}_to_{income_data["period"]["end_date"]}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    def export_balance_sheet_pdf(self, balance_sheet_data):
        """Export Balance Sheet to PDF"""
        # Validate input data
        if not balance_sheet_data:
            raise ValueError("Balance sheet data is required")
        
        if 'as_of_date' not in balance_sheet_data:
            raise ValueError("Balance sheet data must include 'as_of_date'")
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=12,
            alignment=TA_CENTER
        )
        
        user_name = self.user.get_full_name() or self.user.username
        title = Paragraph(f"{user_name}<br/>Balance Sheet", title_style)
        subtitle = Paragraph(f"As of {balance_sheet_data['as_of_date']}", styles['Normal'])
        elements.append(title)
        elements.append(subtitle)
        elements.append(Spacer(1, 0.3*inch))
        
        # Get data with safe defaults
        assets = balance_sheet_data.get('assets', {})
        liabilities = balance_sheet_data.get('liabilities', {})
        equity = balance_sheet_data.get('equity', {})
        
        # Table data
        data = [
            ['Account', 'Amount'],
            ['ASSETS', ''],
            ['Current Assets', f"{float(assets.get('current_assets', 0)):,.2f}"],
            ['Total Assets', f"{float(assets.get('total_assets', 0)):,.2f}"],
            ['', ''],
            ['LIABILITIES', ''],
            ['Current Liabilities', f"{float(liabilities.get('current_liabilities', 0)):,.2f}"],
            ['Total Liabilities', f"{float(liabilities.get('total_liabilities', 0)):,.2f}"],
            ['', ''],
            ['EQUITY', ''],
            ['Retained Earnings', f"{float(equity.get('retained_earnings', 0)):,.2f}"],
            ['Other Equity', f"{float(equity.get('other_equity', 0)):,.2f}"],
            ['Total Equity', f"{float(equity.get('total_equity', 0)):,.2f}"],
            ['', ''],
            ['Total Liabilities & Equity', f"{float(balance_sheet_data.get('total_liabilities_and_equity', 0)):,.2f}"],
        ]
        
        table = Table(data, colWidths=[4*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (0, 1), 'Helvetica-Bold'),
            ('FONTNAME', (0, 5), (0, 5), 'Helvetica-Bold'),
            ('FONTNAME', (0, 9), (0, 9), 'Helvetica-Bold'),
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        buffer.seek(0)
        file_content = buffer.read()
        buffer.close()
        
        # Ensure filename is safe
        as_of_date = balance_sheet_data.get('as_of_date', 'unknown')
        if isinstance(as_of_date, str):
            filename = f"balance_sheet_{as_of_date}.pdf"
        else:
            filename = f"balance_sheet_{as_of_date}.pdf"
        
        response = HttpResponse(file_content, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Content-Length'] = len(file_content)
        return response
    
    def export_income_statement_pdf(self, income_data):
        """Export Income Statement to PDF"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#70AD47'),
            spaceAfter=12,
            alignment=TA_CENTER
        )
        
        title = Paragraph(f"{self.user.get_full_name() or self.user.username}<br/>Income Statement", title_style)
        start_date = income_data.get('period', {}).get('start_date', '')
        end_date = income_data.get('period', {}).get('end_date', '')
        period = f"{start_date} to {end_date}" if start_date or end_date else 'N/A'
        subtitle = Paragraph(f"Period: {period}", styles['Normal'])
        elements.append(title)
        elements.append(subtitle)
        elements.append(Spacer(1, 0.3*inch))
        
        # Table data
        data = [
            ['Description', 'Amount'],
            ['REVENUE', ''],
            ['Total Revenue', f"{float(income_data.get('revenue', 0)):,.2f}"],
            ['', ''],
            ['EXPENSES', ''],
        ]
        
        expenses = income_data.get('expenses', {})
        breakdown = expenses.get('breakdown', {}) if isinstance(expenses, dict) else {}
        for category, amount in breakdown.items():
            try:
                amount_val = float(amount)
            except Exception:
                amount_val = 0.0
            data.append([category, f"{amount_val:,.2f}"])

        total_expenses = expenses.get('total_expenses', 0)
        try:
            total_expenses_val = float(total_expenses)
        except Exception:
            total_expenses_val = 0.0

        data.append(['Total Expenses', f"{total_expenses_val:,.2f}"])
        data.append(['', ''])

        try:
            net_income_val = float(income_data.get('net_income', 0))
        except Exception:
            net_income_val = 0.0
        data.append(['NET INCOME', f"{net_income_val:,.2f}"])

        try:
            profit_margin_val = float(income_data.get('profit_margin', 0))
            profit_margin_str = f"{profit_margin_val:.2f}%"
        except Exception:
            profit_margin_str = "0.00%"
        data.append(['Profit Margin', profit_margin_str])
        
        table = Table(data, colWidths=[4*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#70AD47')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        # Use the sanitized start_date/end_date variables for the filename
        filename = f'income_statement_{start_date}_to_{end_date}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    def export_transactions_excel(self, bills):
        """Export transaction ledger to Excel for auditing"""
        import logging
        logger = logging.getLogger(__name__)
        
        logger.info(f"=== export_transactions_excel called ===")
        logger.info(f"Method: export_transactions_excel (Transaction Ledger)")
        logger.info(f"User: {self.user.username}")
        logger.info(f"Bills count: {len(list(bills)) if hasattr(bills, '__iter__') else 'QuerySet'}")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Transaction Ledger"
        
        # Title
        ws['A1'] = f"{self.user.get_full_name() or self.user.username} - Transaction Ledger"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(size=12, italic=True)
        
        # Headers - Only essential transaction ledger columns
        # Note: OCR text and confidence score are excluded for clean reporting
        headers = ['Date', 'Invoice #', 'Vendor', 'Category', 'Type', 'Debit', 'Credit', 'Balance', 'Notes']
        row = 4
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        running_balance = Decimal('0.00')
        row += 1
        
        bill_count = 0
        for bill in bills:
            bill_count += 1
            logger.info(f"Writing bill {bill_count} to row {row}: {bill.vendor}, Date: {bill.bill_date}, Amount: {bill.amount_npr}")
            
            ws.cell(row=row, column=1).value = str(bill.bill_date) if bill.bill_date else ''
            ws.cell(row=row, column=2).value = bill.invoice_number or ''
            ws.cell(row=row, column=3).value = bill.vendor or ''
            ws.cell(row=row, column=4).value = bill.category.name if bill.category else ''
            ws.cell(row=row, column=5).value = bill.transaction_type
            
            if bill.is_debit:
                ws.cell(row=row, column=6).value = float(bill.amount_npr) if bill.amount_npr else 0
                ws.cell(row=row, column=6).number_format = '#,##0.00'
                ws.cell(row=row, column=7).value = 0
                running_balance -= bill.amount_npr if bill.amount_npr else Decimal('0.00')
            else:
                ws.cell(row=row, column=6).value = 0
                ws.cell(row=row, column=7).value = float(bill.amount_npr) if bill.amount_npr else 0
                ws.cell(row=row, column=7).number_format = '#,##0.00'
                running_balance += bill.amount_npr if bill.amount_npr else Decimal('0.00')
            
            ws.cell(row=row, column=8).value = float(running_balance)
            ws.cell(row=row, column=8).number_format = '#,##0.00'
            ws.cell(row=row, column=9).value = bill.notes or ''
            # Explicitly not including: confidence_score (column 10) or ocr_text (column 11)
            
            row += 1
        
        logger.info(f"Exported {bill_count} bills to Excel for user {self.user.username}")
        
        # Adjust column widths
        for col_num, width in enumerate([12, 15, 25, 20, 10, 12, 12, 12, 30], 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="transaction_ledger_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        return response
    
    def export_cash_flow_excel(self, cash_flow_data):
        """
        Export Cash Flow Statement to Excel
        
        Args:
            cash_flow_data: Dict from NRBFinancialStatements.get_cash_flow_statement()
        
        Returns:
            HttpResponse with Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Cash Flow Statement"
        
        # Title
        ws['A1'] = cash_flow_data.get('company_name', 'Company')
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = cash_flow_data.get('statement_name', 'Cash Flow Statement')
        ws['A2'].font = Font(size=14, bold=True)
        ws['A3'] = f"Period: {cash_flow_data.get('period', 'N/A')}"
        ws['A3'].font = Font(size=12, italic=True)
        ws['A4'] = f"Currency: {cash_flow_data.get('currency', 'NPR')}"
        
        # Headers
        row = 6
        ws[f'A{row}'] = "Particulars"
        ws[f'B{row}'] = "This Year"
        ws[f'C{row}'] = "Last Year"
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{col}{row}'].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            ws[f'{col}{row}'].font = Font(color="FFFFFF", bold=True)
        
        row += 1
        
        # Operating Activities
        ws[f'A{row}'] = "A. Cash Flow from Operating Activities"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        current = cash_flow_data.get('current_year', {}).get('operating_activities', {})
        last = cash_flow_data.get('last_year', {}).get('operating_activities', {})
        
        operating_items = [
            ('Profit Before Tax', 'profit_before_tax'),
            ('Adjustments for:', None),
            ('  Depreciation and Amortization', 'depreciation'),
            ('Operating Profit Before Working Capital Changes', 'operating_profit_before_changes'),
            ('Changes in Operating Assets and Liabilities', None),
            ('  Changes in Operating Assets', 'changes_in_assets'),
            ('  Changes in Operating Liabilities', 'changes_in_liabilities'),
            ('Cash Generated from Operations', 'cash_from_operations'),
            ('Income Tax Paid', 'tax_paid'),
            ('Net Cash from Operating Activities', 'net_cash_from_operating'),
        ]
        
        for label, key in operating_items:
            ws[f'A{row}'] = label
            if key and key in current:
                ws[f'B{row}'] = float(current[key])
                ws[f'B{row}'].number_format = '#,##0.00'
            if key and key in last:
                ws[f'C{row}'] = float(last[key])
                ws[f'C{row}'].number_format = '#,##0.00'
            if key == 'net_cash_from_operating':
                ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        row += 1
        
        # Investing Activities
        ws[f'A{row}'] = "B. Cash Flow from Investing Activities"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        current_inv = cash_flow_data.get('current_year', {}).get('investing_activities', {})
        last_inv = cash_flow_data.get('last_year', {}).get('investing_activities', {})
        
        investing_items = [
            ('Purchase of Property and Equipment', 'purchase_of_property'),
            ('Proceeds from Sale of Assets', 'proceeds_from_sale'),
            ('Net Cash from Investing Activities', 'net_cash_from_investing'),
        ]
        
        for label, key in investing_items:
            ws[f'A{row}'] = label
            if key in current_inv:
                ws[f'B{row}'] = float(current_inv[key])
                ws[f'B{row}'].number_format = '#,##0.00'
            if key in last_inv:
                ws[f'C{row}'] = float(last_inv[key])
                ws[f'C{row}'].number_format = '#,##0.00'
            if key == 'net_cash_from_investing':
                ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        row += 1
        
        # Financing Activities
        ws[f'A{row}'] = "C. Cash Flow from Financing Activities"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        current_fin = cash_flow_data.get('current_year', {}).get('financing_activities', {})
        last_fin = cash_flow_data.get('last_year', {}).get('financing_activities', {})
        
        financing_items = [
            ('Proceeds from Borrowings', 'proceeds_from_borrowing'),
            ('Repayment of Borrowings', 'repayment_of_borrowing'),
            ('Dividends Paid', 'dividends_paid'),
            ('Net Cash from Financing Activities', 'net_cash_from_financing'),
        ]
        
        for label, key in financing_items:
            ws[f'A{row}'] = label
            if key in current_fin:
                ws[f'B{row}'] = float(current_fin[key])
                ws[f'B{row}'].number_format = '#,##0.00'
            if key in last_fin:
                ws[f'C{row}'] = float(last_fin[key])
                ws[f'C{row}'].number_format = '#,##0.00'
            if key == 'net_cash_from_financing':
                ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        row += 1
        
        # Summary
        current_sum = cash_flow_data.get('current_year', {})
        last_sum = cash_flow_data.get('last_year', {})
        
        summary_items = [
            ('Net Increase/(Decrease) in Cash and Cash Equivalents', 'net_change_in_cash'),
            ('Cash and Cash Equivalents at Beginning', 'cash_beginning'),
            ('Cash and Cash Equivalents at End', 'cash_ending'),
        ]
        
        for label, key in summary_items:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            if key in current_sum:
                ws[f'B{row}'] = float(current_sum[key])
                ws[f'B{row}'].number_format = '#,##0.00'
                ws[f'B{row}'].font = Font(bold=True)
            if key in last_sum:
                ws[f'C{row}'] = float(last_sum[key])
                ws[f'C{row}'].number_format = '#,##0.00'
                ws[f'C{row}'].font = Font(bold=True)
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def export_cash_flow_pdf(self, cash_flow_data):
        """
        Export Cash Flow Statement to PDF
        
        Args:
            cash_flow_data: Dict from NRBFinancialStatements.get_cash_flow_statement()
        
        Returns:
            PDF bytes
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        elements.append(Paragraph(cash_flow_data.get('company_name', 'Company'), title_style))
        elements.append(Paragraph(cash_flow_data.get('statement_name', 'Cash Flow Statement'), title_style))
        elements.append(Paragraph(f"Period: {cash_flow_data.get('period', 'N/A')}", styles['Normal']))
        elements.append(Paragraph(f"Currency: {cash_flow_data.get('currency', 'NPR')}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Create table data
        table_data = [
            ['Particulars', 'This Year', 'Last Year']
        ]
        
        current = cash_flow_data.get('current_year', {})
        last = cash_flow_data.get('last_year', {})
        
        # Operating Activities
        table_data.append(['A. Cash Flow from Operating Activities', '', ''])
        
        current_op = current.get('operating_activities', {})
        last_op = last.get('operating_activities', {})
        
        operating_items = [
            ('Profit Before Tax', 'profit_before_tax'),
            ('  Depreciation and Amortization', 'depreciation'),
            ('Operating Profit Before Working Capital Changes', 'operating_profit_before_changes'),
            ('  Changes in Operating Assets', 'changes_in_assets'),
            ('  Changes in Operating Liabilities', 'changes_in_liabilities'),
            ('Cash Generated from Operations', 'cash_from_operations'),
            ('Income Tax Paid', 'tax_paid'),
            ('Net Cash from Operating Activities', 'net_cash_from_operating'),
        ]
        
        for label, key in operating_items:
            current_val = f"NPR {current_op.get(key, 0):,.2f}" if key in current_op else ''
            last_val = f"NPR {last_op.get(key, 0):,.2f}" if key in last_op else ''
            table_data.append([label, current_val, last_val])
        
        table_data.append(['', '', ''])
        
        # Investing Activities
        table_data.append(['B. Cash Flow from Investing Activities', '', ''])
        
        current_inv = current.get('investing_activities', {})
        last_inv = last.get('investing_activities', {})
        
        investing_items = [
            ('Purchase of Property and Equipment', 'purchase_of_property'),
            ('Proceeds from Sale of Assets', 'proceeds_from_sale'),
            ('Net Cash from Investing Activities', 'net_cash_from_investing'),
        ]
        
        for label, key in investing_items:
            current_val = f"NPR {current_inv.get(key, 0):,.2f}" if key in current_inv else ''
            last_val = f"NPR {last_inv.get(key, 0):,.2f}" if key in last_inv else ''
            table_data.append([label, current_val, last_val])
        
        table_data.append(['', '', ''])
        
        # Financing Activities
        table_data.append(['C. Cash Flow from Financing Activities', '', ''])
        
        current_fin = current.get('financing_activities', {})
        last_fin = last.get('financing_activities', {})
        
        financing_items = [
            ('Proceeds from Borrowings', 'proceeds_from_borrowing'),
            ('Repayment of Borrowings', 'repayment_of_borrowing'),
            ('Dividends Paid', 'dividends_paid'),
            ('Net Cash from Financing Activities', 'net_cash_from_financing'),
        ]
        
        for label, key in financing_items:
            current_val = f"NPR {current_fin.get(key, 0):,.2f}" if key in current_fin else ''
            last_val = f"NPR {last_fin.get(key, 0):,.2f}" if key in last_fin else ''
            table_data.append([label, current_val, last_val])
        
        table_data.append(['', '', ''])
        
        # Summary
        summary_items = [
            ('Net Increase/(Decrease) in Cash', 'net_change_in_cash'),
            ('Cash and Cash Equivalents at Beginning', 'cash_beginning'),
            ('Cash and Cash Equivalents at End', 'cash_ending'),
        ]
        
        for label, key in summary_items:
            current_val = f"NPR {current.get(key, 0):,.2f}" if key in current else ''
            last_val = f"NPR {last.get(key, 0):,.2f}" if key in last else ''
            table_data.append([label, current_val, last_val])
        
        # Create table
        table = Table(table_data, colWidths=[4*inch, 1.5*inch, 1.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F81BD')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        buffer.seek(0)
        return buffer.getvalue()
